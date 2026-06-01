# -*- coding: utf-8 -*-
"""
KnowledgeImporter - 知识导入器

导入用户历史资料，快速搭建知识库。

核心功能：
- 支持多种文件格式：PDF、Word、Markdown、TXT、Excel、CSV
- 支持单文件和目录批量导入
- 支持 URL 网页导入（带安全验证）
- 自动调用知识编译器提取知识
- 进度追踪、取消机制、断点续传

安全特性：
- URL 白名单验证，防止 SSRF 攻击
- 文件大小限制，防止内存溢出
- 路径安全检查，防止路径遍历

使用方式：
```python
importer = KnowledgeImporter(knowledge_root="data/users/test/knowledge")

# 导入单个文件
result = importer.import_file("report.pdf", auto_extract=True)

# 批量导入目录
results = importer.import_directory("./documents", auto_extract=True)

# 导入 URL
result = importer.import_url("https://example.com/article")

# 取消导入
importer.cancel_import()
```
"""

__all__ = [
    "KnowledgeImporter",
    "ImportResult",
    "FileParser",
    "ImportProgress"
]

import os
import re
import json
import logging
import hashlib
import threading
import socket
import ipaddress
from pathlib import Path
from typing import Dict, List, Optional, Set, Any, Callable, TYPE_CHECKING
from dataclasses import dataclass, field
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

if TYPE_CHECKING:
    from .compiler import CompiledKnowledge
from urllib.parse import urlparse
import urllib.request
import urllib.error

logger = logging.getLogger(__name__)


# ==================== 安全配置 ====================

# 文件大小限制（50MB）
MAX_FILE_SIZE = 50 * 1024 * 1024

# URL 响应大小限制（10MB）
MAX_URL_SIZE = 10 * 1024 * 1024

# 允许的 URL 协议
ALLOWED_PROTOCOLS = {'http', 'https'}

# 内网 IP 黑名单（SSRF 防护）
PRIVATE_IP_RANGES = [
    ipaddress.ip_network('127.0.0.0/8'),      # Loopback
    ipaddress.ip_network('10.0.0.0/8'),       # Class A private
    ipaddress.ip_network('172.16.0.0/12'),    # Class B private
    ipaddress.ip_network('192.168.0.0/16'),   # Class C private
    ipaddress.ip_network('169.254.0.0/16'),   # Link-local
    ipaddress.ip_network('::1/128'),          # IPv6 loopback
    ipaddress.ip_network('fc00::/7'),         # IPv6 unique local
    ipaddress.ip_network('fe80::/10'),        # IPv6 link-local
]


# ==================== 数据类 ====================

@dataclass
class ImportResult:
    """
    导入结果
    
    Attributes:
        file_path: 文件路径
        status: 状态（success/failed/skipped/partial）
        content: 解析后的内容
        pages_created: 创建的知识页数量
        entities_extracted: 提取的实体数量
        error_message: 错误信息
        import_time: 导入时间
        file_size: 文件大小（字节）
        compiled_knowledge: 编译结果（由 import_file/import_url 赋值）
    """
    file_path: str
    status: str  # success, failed, skipped, partial
    content: Optional[str] = None
    pages_created: int = 0
    entities_extracted: int = 0
    error_message: str = ""
    import_time: datetime = field(default_factory=datetime.now)
    file_size: int = 0
    compiled_knowledge: Optional["CompiledKnowledge"] = None
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "file_path": self.file_path,
            "status": self.status,
            "pages_created": self.pages_created,
            "entities_extracted": self.entities_extracted,
            "error_message": self.error_message,
            "import_time": self.import_time.isoformat(),
            "file_size": self.file_size
        }


@dataclass
class ImportProgress:
    """
    导入进度
    
    Attributes:
        total_files: 总文件数
        processed_files: 已处理文件数
        failed_files: 失败文件数
        partial_files: 部分成功文件数
        current_file: 当前处理的文件
        start_time: 开始时间
        total_bytes: 已处理的总字节数
        is_cancelled: 是否已取消
    """
    total_files: int = 0
    processed_files: int = 0
    failed_files: int = 0
    partial_files: int = 0
    current_file: str = ""
    start_time: datetime = field(default_factory=datetime.now)
    total_bytes: int = 0
    is_cancelled: bool = False
    
    def update(self, success: bool, file_path: str = "", file_size: int = 0, is_partial: bool = False):
        """更新进度"""
        self.processed_files += 1
        if not success:
            self.failed_files += 1
        elif is_partial:
            self.partial_files += 1
        self.current_file = file_path
        self.total_bytes += file_size
    
    def get_percentage(self) -> float:
        """获取进度百分比"""
        if self.total_files == 0:
            return 0.0
        return (self.processed_files / self.total_files) * 100
    
    def get_elapsed_time(self) -> float:
        """获取已用时间（秒）"""
        return (datetime.now() - self.start_time).total_seconds()
    
    def get_estimated_remaining(self) -> float:
        """获取预计剩余时间（秒）"""
        if self.processed_files == 0:
            return 0.0
        avg_time = self.get_elapsed_time() / self.processed_files
        remaining_files = self.total_files - self.processed_files
        return avg_time * remaining_files
    
    def get_speed(self) -> float:
        """获取处理速度（文件/秒）"""
        elapsed = self.get_elapsed_time()
        if elapsed == 0:
            return 0.0
        return self.processed_files / elapsed


# ==================== 安全工具函数 ====================

def validate_url(url: str) -> str:
    """
    验证 URL 安全性，防止 SSRF 攻击
    
    Args:
        url: 要验证的 URL
    
    Returns:
        验证通过返回规范化 URL
    
    Raises:
        ValueError: URL 不安全或格式错误
    """
    try:
        parsed = urlparse(url)
    except Exception as e:
        raise ValueError(f"Invalid URL format: {e}")
    
    # 检查协议
    if parsed.scheme.lower() not in ALLOWED_PROTOCOLS:
        raise ValueError(
            f"Protocol '{parsed.scheme}' not allowed. Only {ALLOWED_PROTOCOLS} are permitted."
        )
    
    # 检查主机名
    hostname = parsed.hostname
    if not hostname:
        raise ValueError("URL must have a hostname")
    
    # 解析 IP 地址
    try:
        # 尝试直接解析为 IP
        ip = ipaddress.ip_address(hostname)
        if _is_private_ip(ip):
            raise ValueError(f"Private IP address not allowed: {ip}")
    except ValueError:
        # 不是 IP，进行 DNS 解析
        try:
            addr_info = socket.getaddrinfo(hostname, parsed.port or 80)
            for family, _, _, _, sockaddr in addr_info:
                ip = ipaddress.ip_address(sockaddr[0])
                if _is_private_ip(ip):
                    raise ValueError(
                        f"Hostname resolves to private IP: {hostname} -> {ip}"
                    )
        except socket.gaierror as e:
            raise ValueError(f"Cannot resolve hostname: {hostname} ({e})")
    
    # 构建规范化 URL
    return parsed.geturl()


def _is_private_ip(ip: ipaddress.IPv4Address | ipaddress.IPv6Address) -> bool:
    """检查 IP 是否为内网地址"""
    for network in PRIVATE_IP_RANGES:
        if ip in network:
            return True
    return False


def sanitize_path(path: str, base_dir: Path) -> Path:
    """
    清理路径，防止路径遍历攻击
    
    Args:
        path: 输入路径
        base_dir: 基础目录
    
    Returns:
        安全的绝对路径
    
    Raises:
        ValueError: 路径不安全
    """
    # 转换为 Path 对象
    input_path = Path(path)
    
    # 如果是相对路径，基于 base_dir 解析
    if not input_path.is_absolute():
        resolved = (base_dir / input_path).resolve()
    else:
        resolved = input_path.resolve()
    
    # 检查是否在允许的目录内
    try:
        resolved.relative_to(base_dir.resolve())
    except ValueError:
        raise ValueError(f"Path traversal detected: {path}")
    
    return resolved


# ==================== 文件解析器 ====================

class FileParser:
    """
    文件解析器
    
    支持的格式：
    - Markdown (.md)
    - 纯文本 (.txt)
    - CSV (.csv)
    - JSON (.json)
    - PDF (.pdf) - 可选，需要 PyPDF2
    - Word (.docx) - 可选，需要 python-docx
    - Excel (.xlsx/.xls) - 可选，需要 openpyxl
    """
    
    # 支持的文件扩展名
    SUPPORTED_EXTENSIONS = {
        'text': ['.txt', '.md', '.markdown', '.rst'],
        'data': ['.csv', '.json'],
        'document': ['.pdf', '.docx', '.doc'],
        'spreadsheet': ['.xlsx', '.xls']
    }
    
    def __init__(self, max_file_size: int = MAX_FILE_SIZE):
        """
        初始化解析器
        
        Args:
            max_file_size: 最大文件大小（字节）
        """
        self.max_file_size = max_file_size
        self._check_optional_dependencies()
    
    def _check_optional_dependencies(self):
        """检查可选依赖"""
        self.has_pdf = self._try_import('PyPDF2')
        self.has_docx = self._try_import('docx')
        self.has_xlsx = self._try_import('openpyxl')
    
    def _try_import(self, module_name: str) -> bool:
        """尝试导入模块"""
        try:
            __import__(module_name)
            return True
        except ImportError:
            return False
    
    def parse_file(self, file_path: str) -> Optional[str]:
        """
        解析文件
        
        Args:
            file_path: 文件路径
        
        Returns:
            解析后的文本内容，失败返回 None
        """
        path = Path(file_path)
        
        if not path.exists():
            logger.error(f"File not found: {file_path}")
            return None
        
        # 检查文件大小
        file_size = path.stat().st_size
        if file_size > self.max_file_size:
            logger.error(
                f"File too large: {file_path} "
                f"({file_size / 1024 / 1024:.1f}MB > {self.max_file_size / 1024 / 1024:.1f}MB)"
            )
            return None
        
        ext = path.suffix.lower()
        
        try:
            # 文本格式
            if ext in self.SUPPORTED_EXTENSIONS['text']:
                return self._parse_text(file_path)
            
            # 数据格式
            if ext in self.SUPPORTED_EXTENSIONS['data']:
                return self._parse_data(file_path)
            
            # PDF
            if ext == '.pdf':
                return self._parse_pdf(file_path)
            
            # Word
            if ext in ['.docx', '.doc']:
                return self._parse_word(file_path)
            
            # Excel
            if ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_path)
            
            # 未知格式，尝试作为文本读取
            logger.warning(f"Unknown format {ext}, trying as text")
            return self._parse_text(file_path)
            
        except Exception as e:
            logger.error(f"Failed to parse {file_path}: {e}")
            return None
    
    def get_file_size(self, file_path: str) -> int:
        """获取文件大小"""
        try:
            return Path(file_path).stat().st_size
        except Exception:
            return 0
    
    def _parse_text(self, file_path: str) -> str:
        """解析文本文件"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        
        # 所有编码都失败
        logger.error(f"Failed to decode {file_path}")
        return ""
    
    def _parse_data(self, file_path: str) -> str:
        """解析数据文件（CSV、JSON）"""
        ext = Path(file_path).suffix.lower()
        
        if ext == '.json':
            return self._parse_json(file_path)
        elif ext == '.csv':
            return self._parse_csv(file_path)
        
        return ""
    
    def _parse_json(self, file_path: str) -> str:
        """解析 JSON 文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 将 JSON 转换为可读文本
            return self._json_to_text(data)
        except Exception as e:
            logger.error(f"Failed to parse JSON {file_path}: {e}")
            return ""
    
    def _json_to_text(self, data: Any, indent: int = 0) -> str:
        """将 JSON 数据转换为文本"""
        lines = []
        prefix = "  " * indent
        
        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    lines.append(f"{prefix}{key}:")
                    lines.append(self._json_to_text(value, indent + 1))
                else:
                    lines.append(f"{prefix}{key}: {value}")
        elif isinstance(data, list):
            for i, item in enumerate(data):
                if isinstance(item, (dict, list)):
                    lines.append(f"{prefix}- [{i}]")
                    lines.append(self._json_to_text(item, indent + 1))
                else:
                    lines.append(f"{prefix}- {item}")
        else:
            lines.append(f"{prefix}{data}")
        
        return "\n".join(lines)
    
    def _parse_csv(self, file_path: str) -> str:
        """
        解析 CSV 文件
        
        使用 csv 标准库正确处理引号内逗号和多行字段。
        """
        import csv
        lines = []
        
        try:
            # 尝试多种编码
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
            content = None
            used_encoding = None
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                    used_encoding = encoding
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                logger.error(f"Failed to decode CSV file: {file_path}")
                return ""
            
            # 使用 csv 标准库解析
            reader = csv.reader(content.strip().splitlines())
            rows = list(reader)
            
            if not rows:
                return ""
            
            # 第一行是表头
            headers = rows[0]
            headers = [h.strip() for h in headers if h.strip()]
            
            lines.append(f"CSV 数据（列：{', '.join(headers)}）\n")
            
            # 处理数据行
            for row in rows[1:]:
                if not any(row):  # 跳过空行
                    continue
                    
                for h, v in zip(headers, row):
                    v = v.strip()
                    if v:
                        lines.append(f"{h}: {v}")
                lines.append("")
            
            return "\n".join(lines)
            
        except Exception as e:
            logger.error(f"Failed to parse CSV {file_path}: {e}")
            return ""
    
    def _parse_pdf(self, file_path: str) -> str:
        """解析 PDF 文件"""
        if not self.has_pdf:
            logger.warning("PyPDF2 not installed, cannot parse PDF")
            return ""
        
        try:
            import PyPDF2
            
            text_parts = []
            
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
            
            return "\n\n".join(text_parts)
            
        except Exception as e:
            logger.error(f"Failed to parse PDF {file_path}: {e}")
            return ""
    
    def _parse_word(self, file_path: str) -> str:
        """解析 Word 文件"""
        if not self.has_docx:
            logger.warning("python-docx not installed, cannot parse Word")
            return ""
        
        try:
            import docx
            
            doc = docx.Document(file_path)
            
            paragraphs = []
            for para in doc.paragraphs:
                if para.text.strip():
                    paragraphs.append(para.text)
            
            return "\n\n".join(paragraphs)
            
        except Exception as e:
            logger.error(f"Failed to parse Word {file_path}: {e}")
            return ""
    
    def _parse_excel(self, file_path: str) -> str:
        """解析 Excel 文件"""
        if not self.has_xlsx:
            logger.warning("openpyxl not installed, cannot parse Excel")
            return ""
        
        wb = None
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            lines = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                lines.append(f"## 工作表: {sheet_name}\n")
                
                for row in sheet.iter_rows(values_only=True):
                    # 过滤空行
                    values = [str(v) if v is not None else "" for v in row]
                    if any(values):
                        lines.append(" | ".join(values))
                
                lines.append("")
            
            return "\n".join(lines)
            
        except Exception as e:
            logger.error(f"Failed to parse Excel {file_path}: {e}")
            return ""
        finally:
            # 确保 workbook 被关闭
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
    
    def get_supported_extensions(self) -> List[str]:
        """获取所有支持的扩展名"""
        extensions = []
        for ext_list in self.SUPPORTED_EXTENSIONS.values():
            extensions.extend(ext_list)
        return extensions


# ==================== 知识导入器 ====================

class KnowledgeImporter:
    """
    知识导入器
    
    导入用户历史资料并自动提取知识。
    
    Attributes:
        knowledge_root: 知识库根目录
        parser: 文件解析器
        compiler: 知识编译器
    """
    
    def __init__(
        self,
        knowledge_root: Path,
        user_id: Optional[str] = None,
        max_file_size: int = MAX_FILE_SIZE
    ):
        """
        初始化导入器
        
        Args:
            knowledge_root: 知识库根目录
            user_id: 用户ID（可选）
            max_file_size: 最大文件大小（字节）
        """
        self.knowledge_root = Path(knowledge_root)
        self.user_id = user_id
        self.max_file_size = max_file_size
        
        # 创建目录结构
        self.knowledge_root.mkdir(parents=True, exist_ok=True)
        (self.knowledge_root / "imported").mkdir(parents=True, exist_ok=True)
        
        # 初始化组件
        self.parser = FileParser(max_file_size=max_file_size)
        
        # 延迟导入避免循环依赖
        self._compiler = None
        
        # 取消标志
        self._cancel_event = threading.Event()
        
        # 导入统计
        self._stats = {
            "total_imported": 0,
            "total_failed": 0,
            "total_partial": 0,
            "total_pages_created": 0,
            "total_bytes_processed": 0,
            "last_import": None,
            "last_cancel": None
        }
        
        # 导入清单（断点续传）
        # manifest 结构: {file_hash: {"path": str, "imported_at": str, "pages_created": int}}
        self._import_manifest: Dict[str, Dict[str, Any]] = {}
        self._load_import_manifest()
        
        logger.info(f"KnowledgeImporter initialized: knowledge_root={knowledge_root}")
    
    @property
    def compiler(self):
        """延迟加载知识编译器"""
        if self._compiler is None:
            from .compiler import KnowledgeCompiler
            self._compiler = KnowledgeCompiler(knowledge_root=self.knowledge_root)
        return self._compiler
    
    def cancel_import(self):
        """取消正在进行的导入"""
        self._cancel_event.set()
        self._stats["last_cancel"] = datetime.now().isoformat()
        logger.info("Import cancellation requested")
    
    def _check_cancelled(self) -> bool:
        """检查是否已取消"""
        if self._cancel_event.is_set():
            logger.info("Import cancelled by user")
            return True
        return False
    
    def _load_import_manifest(self):
        """加载导入清单（断点续传）"""
        manifest_path = self.knowledge_root / ".import_manifest.json"
        if manifest_path.exists():
            try:
                self._import_manifest = json.loads(manifest_path.read_text(encoding='utf-8'))
            except Exception as e:
                logger.warning(f"Failed to load import manifest: {e}")
                self._import_manifest = {}
    
    def _save_import_manifest(self):
        """保存导入清单"""
        manifest_path = self.knowledge_root / ".import_manifest.json"
        try:
            manifest_path.write_text(
                json.dumps(self._import_manifest, ensure_ascii=False, indent=2),
                encoding='utf-8'
            )
        except Exception as e:
            logger.warning(f"Failed to save import manifest: {e}")
    
    def _compute_file_hash(self, file_path: str) -> str:
        """计算文件哈希"""
        try:
            hasher = hashlib.md5(usedforsecurity=False)
            with open(file_path, 'rb') as f:
                for chunk in iter(lambda: f.read(8192), b''):
                    hasher.update(chunk)
            return hasher.hexdigest()
        except Exception:
            return ""
    
    def import_file(
        self,
        file_path: str,
        auto_extract: bool = True,
        source_info: Optional[Dict] = None,
        skip_if_imported: bool = True
    ) -> ImportResult:
        """
        导入单个文件
        
        Args:
            file_path: 文件路径
            auto_extract: 是否自动提取知识
            source_info: 来源信息
            skip_if_imported: 是否跳过已导入的文件
        
        Returns:
            ImportResult: 导入结果
        """
        # 检查取消
        if self._check_cancelled():
            return ImportResult(
                file_path=file_path,
                status="failed",
                error_message="Import cancelled"
            )
        
        logger.info(f"Importing file: {file_path}")
        
        # 检查是否已导入（断点续传）
        file_hash = self._compute_file_hash(file_path)
        if skip_if_imported and file_hash and file_hash in self._import_manifest:
            logger.info(f"File already imported, skipping: {file_path}")
            return ImportResult(
                file_path=file_path,
                status="skipped",
                error_message="Already imported"
            )
        
        # 安全检查路径（统一使用 sanitize_path）
        try:
            safe_path = Path(file_path)
            # 使用 resolve() 解析所有符号链接和路径遍历
            resolved_path = safe_path.resolve()
            
            # 检查路径遍历攻击
            # 1. 检查解析后的路径是否仍包含 ".."（异常情况）
            if ".." in str(resolved_path):
                return ImportResult(
                    file_path=file_path,
                    status="failed",
                    error_message="Path traversal detected in resolved path"
                )
            
            # 2. 对于相对路径，确保在允许的目录内
            if not safe_path.is_absolute():
                safe_path = sanitize_path(file_path, Path.cwd())
            else:
                # 绝对路径也需要验证：检查是否指向敏感系统目录
                sensitive_dirs = [
                    Path("/etc"), Path("/root"), Path("/var/log"),
                    Path("C:\\Windows\\System32"), Path("C:\\Users\\Administrator"),
                ]
                for sensitive_dir in sensitive_dirs:
                    try:
                        if sensitive_dir.exists():
                            resolved_path.relative_to(sensitive_dir.resolve())
                            return ImportResult(
                                file_path=file_path,
                                status="failed",
                                error_message=f"Access to sensitive directory denied: {sensitive_dir}"
                            )
                    except ValueError:
                        pass  # 不在敏感目录内，继续检查
                
                safe_path = resolved_path
        except ValueError as e:
            return ImportResult(
                file_path=file_path,
                status="failed",
                error_message=str(e)
            )
        
        # 解析文件
        content = self.parser.parse_file(str(safe_path))
        
        if content is None:
            return ImportResult(
                file_path=file_path,
                status="failed",
                error_message="Failed to parse file"
            )
        
        if not content.strip():
            return ImportResult(
                file_path=file_path,
                status="skipped",
                error_message="Empty content"
            )
        
        # 创建结果
        file_size = self.parser.get_file_size(str(safe_path))
        result = ImportResult(
            file_path=file_path,
            status="success",
            content=content,
            file_size=file_size
        )
        
        # 自动提取知识
        if auto_extract:
            try:
                knowledge = self.compiler.compile_research(
                    raw_content=content,
                    source_info=source_info or {
                        "title": Path(file_path).stem,
                        "type": "imported_file",
                        "path": file_path
                    }
                )
                
                # 先保存编译结果引用，再写文件系统
                # 即使 save_knowledge 失败，compiled_knowledge 也不会丢失
                result.compiled_knowledge = knowledge
                
                # 保存知识页
                self.compiler.save_knowledge(knowledge)
                
                # 更新结果
                stats = knowledge.get_stats()
                result.pages_created = stats["total"]
                result.entities_extracted = stats["entities"]
                
                logger.info(f"Extracted {stats['total']} pages from {file_path}")
                
            except Exception as e:
                logger.error(f"Failed to extract knowledge: {e}")
                # 标记为部分成功
                result.status = "partial"
                result.error_message = f"Knowledge extraction failed: {e}"
        
        # 更新统计
        self._stats["total_imported"] += 1
        if result.status == "partial":
            self._stats["total_partial"] += 1
        self._stats["total_pages_created"] += result.pages_created
        self._stats["total_bytes_processed"] += file_size
        self._stats["last_import"] = datetime.now().isoformat()
        
        # 保存到导入清单
        if file_hash:
            self._import_manifest[file_hash] = {
                "path": file_path,
                "imported_at": datetime.now().isoformat(),
                "pages_created": result.pages_created
            }
            self._save_import_manifest()
        
        # 保存原始文件副本
        self._save_imported_file(file_path, content)
        
        return result
    
    def import_directory(
        self,
        directory_path: str,
        auto_extract: bool = True,
        recursive: bool = True,
        progress_callback: Optional[Callable[[ImportProgress], None]] = None,
        max_workers: int = 4,
        skip_if_imported: bool = True
    ) -> List[ImportResult]:
        """
        批量导入目录
        
        Args:
            directory_path: 目录路径
            auto_extract: 是否自动提取知识
            recursive: 是否递归子目录
            progress_callback: 进度回调函数
            max_workers: 最大并发数
            skip_if_imported: 是否跳过已导入的文件
        
        Returns:
            导入结果列表
        """
        logger.info(f"Importing directory: {directory_path}")
        
        # 重置取消标志
        self._cancel_event.clear()
        
        # 收集文件
        files = self._collect_files(directory_path, recursive)
        
        if not files:
            logger.warning(f"No supported files found in {directory_path}")
            return []
        
        # 初始化进度
        progress = ImportProgress(total_files=len(files))
        
        results = []
        
        # 并发处理文件
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {
                executor.submit(
                    self.import_file,
                    str(file),
                    auto_extract,
                    None,
                    skip_if_imported
                ): file for file in files
            }
            
            for future in as_completed(future_to_file):
                # 检查取消
                if self._check_cancelled():
                    # 取消剩余任务
                    for f in future_to_file:
                        f.cancel()
                    progress.is_cancelled = True
                    break
                
                file = future_to_file[future]
                
                try:
                    result = future.result()
                    results.append(result)
                    
                    progress.update(
                        success=result.status in ("success", "partial"),
                        file_path=str(file),
                        file_size=result.file_size,
                        is_partial=result.status == "partial"
                    )
                    
                except Exception as e:
                    logger.error(f"Failed to import {file}: {e}")
                    results.append(ImportResult(
                        file_path=str(file),
                        status="failed",
                        error_message=str(e)
                    ))
                    progress.update(success=False, file_path=str(file))
                
                # 回调进度
                if progress_callback:
                    progress_callback(progress)
        
        # 汇总统计
        success_count = sum(1 for r in results if r.status == "success")
        partial_count = sum(1 for r in results if r.status == "partial")
        failed_count = sum(1 for r in results if r.status == "failed")
        skipped_count = sum(1 for r in results if r.status == "skipped")
        
        logger.info(
            f"Directory import complete: "
            f"{success_count} success, {partial_count} partial, "
            f"{failed_count} failed, {skipped_count} skipped"
        )
        
        return results
    
    def _collect_files(
        self,
        directory_path: str,
        recursive: bool = True
    ) -> List[Path]:
        """收集目录中的文件"""
        directory = Path(directory_path)
        supported_ext = set(self.parser.get_supported_extensions())
        
        files = []
        
        if recursive:
            pattern = "**/*"
        else:
            pattern = "*"
        
        for file in directory.glob(pattern):
            if file.is_file() and file.suffix.lower() in supported_ext:
                files.append(file)
        
        return sorted(files)
    
    def import_url(
        self,
        url: str,
        auto_extract: bool = True,
        timeout: int = 30,
        max_size: int = MAX_URL_SIZE,
        retries: int = 3
    ) -> ImportResult:
        """
        导入 URL 内容
        
        Args:
            url: 网页 URL
            auto_extract: 是否自动提取知识
            timeout: 超时时间（秒）
            max_size: 最大响应大小（字节）
            retries: 重试次数
        
        Returns:
            ImportResult: 导入结果
        """
        logger.info(f"Importing URL: {url}")
        
        # 验证 URL 安全性
        try:
            safe_url = validate_url(url)
        except ValueError as e:
            logger.error(f"URL validation failed: {e}")
            return ImportResult(
                file_path=url,
                status="failed",
                error_message=f"Security violation: {e}"
            )
        
        # 带重试的请求
        last_error = None
        html_content = None
        for attempt in range(retries):
            try:
                
                # 请求网页
                req = urllib.request.Request(
                    safe_url,
                    headers={'User-Agent': 'Mozilla/5.0 (compatible; KnowledgeImporter/1.0)'}
                )
                
                with urllib.request.urlopen(req, timeout=timeout) as response:
                    # 检查响应大小
                    content_length = response.headers.get('Content-Length')
                    if content_length and int(content_length) > max_size:
                        return ImportResult(
                            file_path=url,
                            status="failed",
                            error_message=f"Response too large: {content_length} bytes"
                        )
                    
                    # 读取内容（限制大小）
                    html_content = response.read(max_size + 1)
                    if len(html_content) > max_size:
                        return ImportResult(
                            file_path=url,
                            status="failed",
                            error_message=f"Response exceeds max size: {max_size} bytes"
                        )
                    
                    html_content = html_content.decode('utf-8', errors='ignore')
                
                # 成功，跳出重试循环
                break
                
            except urllib.error.HTTPError as e:
                last_error = f"HTTP {e.code}: {e.reason}"
                logger.warning(f"URL fetch attempt {attempt + 1} failed: {last_error}")
            except urllib.error.URLError as e:
                last_error = f"URL error: {e.reason}"
                logger.warning(f"URL fetch attempt {attempt + 1} failed: {last_error}")
            except Exception as e:
                last_error = str(e)
                logger.warning(f"URL fetch attempt {attempt + 1} failed: {e}")
        
        # 所有重试都失败
        if html_content is None or last_error:
            return ImportResult(
                file_path=url,
                status="failed",
                error_message=f"Failed after {retries} retries: {last_error or 'No content retrieved'}"
            )
        
        # 提取文本
        text = self._extract_text_from_html(html_content)
        
        if not text.strip():
            return ImportResult(
                file_path=url,
                status="failed",
                error_message="No text content extracted"
            )
        
        # 创建结果
        file_size = len(text.encode('utf-8'))
        result = ImportResult(
            file_path=url,
            status="success",
            content=text,
            file_size=file_size
        )
        
        # 自动提取知识
        if auto_extract:
            try:
                knowledge = self.compiler.compile_research(
                    raw_content=text,
                    source_info={
                        "title": url,
                        "type": "url",
                        "url": url
                    }
                )
                
                result.compiled_knowledge = knowledge
                
                self.compiler.save_knowledge(knowledge)
                
                stats = knowledge.get_stats()
                result.pages_created = stats["total"]
                result.entities_extracted = stats["entities"]
                
            except Exception as e:
                logger.error(f"Failed to extract knowledge from URL: {e}")
                result.status = "partial"
                result.error_message = f"Knowledge extraction failed: {e}"
        
        # 更新统计
        self._stats["total_imported"] += 1
        if result.status == "partial":
            self._stats["total_partial"] += 1
        self._stats["total_pages_created"] += result.pages_created
        self._stats["total_bytes_processed"] += file_size
        self._stats["last_import"] = datetime.now().isoformat()
        
        return result
    
    def _extract_text_from_html(self, html: str) -> str:
        """从 HTML 提取文本"""
        # 移除脚本和样式
        html = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
        html = re.sub(r'<style[^>]*>.*?</style>', '', html, flags=re.DOTALL | re.IGNORECASE)
        html = re.sub(r'<!--.*?-->', '', html, flags=re.DOTALL)  # 注释
        
        # 移除标签
        text = re.sub(r'<[^>]+>', ' ', html)
        
        # 清理空白
        text = re.sub(r'\s+', ' ', text)
        
        # 提取标题
        title_match = re.search(r'<title>(.*?)</title>', html, re.IGNORECASE)
        if title_match:
            title = title_match.group(1).strip()
            text = f"# {title}\n\n{text}"
        
        return text.strip()
    
    def _save_imported_file(self, file_path: str, content: str):
        """保存导入的文件副本"""
        try:
            # 生成唯一文件名
            file_hash = hashlib.md5(file_path.encode(), usedforsecurity=False).hexdigest()[:8]
            safe_name = re.sub(r'[^\w\u4e00-\u9fa5.-]', '_', Path(file_path).stem)
            file_name = f"{safe_name}_{file_hash}.txt"
            
            save_path = self.knowledge_root / "imported" / file_name
            
            with open(save_path, 'w', encoding='utf-8') as f:
                f.write(f"Source: {file_path}\n")
                f.write(f"Imported: {datetime.now().isoformat()}\n")
                f.write("-" * 50 + "\n\n")
                f.write(content)
            
        except Exception as e:
            logger.warning(f"Failed to save imported file: {e}")
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的文件格式"""
        return self.parser.get_supported_extensions()
    
    def get_stats(self) -> Dict[str, Any]:
        """
        获取导入统计
        
        Returns:
            包含详细统计信息的字典
        """
        stats = self._stats.copy()
        
        # 计算衍生统计
        if stats["total_imported"] > 0:
            stats["failure_rate"] = stats["total_failed"] / stats["total_imported"]
            stats["partial_rate"] = stats["total_partial"] / stats["total_imported"]
        else:
            stats["failure_rate"] = 0.0
            stats["partial_rate"] = 0.0
        
        # 格式化字节大小
        stats["total_mb_processed"] = stats["total_bytes_processed"] / (1024 * 1024)
        
        return stats
    
    def clear_manifest(self):
        """清除导入清单"""
        self._import_manifest = {}
        manifest_path = self.knowledge_root / ".import_manifest.json"
        if manifest_path.exists():
            manifest_path.unlink()
        logger.info("Import manifest cleared")